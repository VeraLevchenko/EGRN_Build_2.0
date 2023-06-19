import openpyxl

# Программа возвращает правообладателя здания из выписки ЕГРН и вид права
# Выписки хранятся в папке ЕГРН текущего каталога
# В результате создает таблицу в указанной папке и открывает ее.
# Исходная таблица должна быть .xlsx и содержать 2 колонки:
# id	Кадастровыйномер    Видправа    Землепользователь
# ---------------------------------------------------------
# 12    42:30:0501005:113   аренда       Иванов Иван Иванович
#
import os
import random
import openpyxl
import time
import xml.etree.ElementTree as ET
import datetime

# функция возвращает список .xml файлов из папки, включая подпапки, кроме proto_.xml
def get_file_list(path_Data):
    filelist = []
    for root, dirs, files in os.walk(path_Data):
        for file in files:
            if file.endswith(".xml") and file != "proto_.xml":
                filelist.append(os.path.join(root, file))
    return filelist

def getFilename(cadnum):
    file_egrn = []
    path_Data = os.path.abspath(os.curdir) + "\ЕГРН"
    filelist = get_file_list(path_Data)
    cadnum = cadnum.replace(":", "_")
    for file in filelist:
        if cadnum in file:
            file_egrn.append(file)
    return file_egrn

def getRight_holders(file_name):
    tree = ET.parse(str(file_name[0]))
    list_right_holders = []
    list_right_types = []
    right_holders_individual = tree.findall('right_records/right_record/right_holders/right_holder/individual')
    for right_holder in right_holders_individual:
        #  ---------------------------возвращает ФИО правообладателя----------------------------
        surname = right_holder.find('surname').text
        name = right_holder.find('name').text
        if right_holder.findall('patronymic'):
            patronymic = right_holder.find('patronymic').text
        else:
            patronymic = "Отчество отсутствует"

        birth_date = right_holder.find('birth_date').text
        birth_place = right_holder.find('birth_place').text

        if right_holder.findall('snils'):
            snils = right_holder.find('snils').text
        else:
            snils = "snils None"

        document_series = right_holder.find('identity_doc/document_series').text
        document_number = right_holder.find('identity_doc/document_number').text
        document_date = right_holder.find('identity_doc/document_date').text
        document_issuer = right_holder.find('identity_doc/document_issuer').text


        list_right_holders.append(surname)
        list_right_holders.append(name)
        list_right_holders.append(patronymic)
        list_right_holders.append(birth_date)
        list_right_holders.append("г.р.")
        list_right_holders.append("место рождения: ")
        list_right_holders.append(birth_place)
        list_right_holders.append("снилс:")
        list_right_holders.append(snils)
        list_right_holders.append("паспорт:")
        list_right_holders.append(document_series)
        list_right_holders.append(document_number)
        list_right_holders.append("выдан ")
        list_right_holders.append(document_date)
        list_right_holders.append(document_issuer)

        print(list_right_holders)


    right_holders_public_formation = tree.findall('right_records/right_record/right_holders/right_holder/public_formation/public_formation_type/municipality')
    for right_holder in right_holders_public_formation:
        #  ---------------------------возвращает Публичного правообладателя----------------------------
        name_pub = right_holder.find('name').text
        list_right_holders.append(name_pub)

    right_holders_public_formation = tree.findall('right_records/right_record/right_data/right_type')
    for right_holder in right_holders_public_formation:
        right_type = right_holder.find('value').text
        list_right_types.append(right_type)

    return list_right_holders, list_right_types

if __name__ == '__main__':

    start = time.time() ## точка отсчета времени
    name_file = "результат" + str(random.randint(1, 10000)) + ".xlsx" ##  формируем имя результирующей таблицы

    wb = openpyxl.load_workbook('src.xlsx')
    sheet = wb.active

    # Создаем объект результирующей таблицы Excel и лист
    table_output = openpyxl.Workbook()
    sheet_table_output = table_output.active

    # Создаем заголовки в результирующей таблице
    sheet_table_output.append(('id', 'Тип', 'Наименование улицы', 'Дом', 'КадастровыйНомерОКС', "ВидПрава", "Правообладатель"))
    max_row = sheet.max_row
    current_date = datetime.datetime.now()

    for i in range(2, max_row):
        message = ""
        list_right_holders = []
        list_right_types = []
        id = sheet[i][0].value
        type_street = str(sheet[i][1].value)
        street = str(sheet[i][2].value)
        house = str(sheet[i][3].value)
        cadnum = str(sheet[i][4].value)
        print(cadnum)
        file_name = getFilename(cadnum)
        print('file_name', file_name)
        if file_name:
            list_right_holders, list_right_types = getRight_holders(file_name)
        else:
            message = "файл ЕГРН отсутствует"
        list_right_holders = ' '.join(str(list_right_holder) for list_right_holder in list_right_holders) # Преобразуем его в строку, элементы разделяем "; "
        list_right_types = '; '.join(str(list_right_type) for list_right_type in list_right_types)
        print(list_right_holders)
        print(list_right_types)
        # Заполняем строку данными
        object = []
        object.append(id)
        object.append(type_street)
        object.append(street)
        object.append(house)
        object.append(cadnum)
        if list_right_holders:
            object.append(list_right_types)
        else:
            object.append(message)
        if list_right_holders:
            object.append(list_right_holders)
        else:
            object.append(message)
        print(object)


        # Добавляем в результирующую таблицу и сохраняем ее
        sheet_table_output.append(object)
        table_output.save(name_file)

    end = time.time() - start #  время работы программы
    os.startfile(name_file) # Открываем результирующий файл
    print("Время работы программы, сек.:", end) # вывод времени




